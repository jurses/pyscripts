#!/usr/bin/env python3
'''
Copyright 2018 Raúl Ulises Martín Hernández

Se concede permiso por la presente, libre de cargos, a cualquier persona que obtenga una copia de este software y de los archivos de documentación asociados (el "Software"), a utilizar el Software sin restricción, incluyendo sin limitación los derechos a usar, copiar, modificar, fusionar, publicar, distribuir, sublicenciar, y/o vender copias del Software, y a permitir a las personas a las que se les proporcione el Software a hacer lo mismo, sujeto a las siguientes condiciones:

El aviso de copyright anterior y este aviso de permiso se incluirán en todas las copias o partes sustanciales del Software.

EL SOFTWARE SE PROPORCIONA "COMO ESTÁ", SIN GARANTÍA DE NINGÚN TIPO, EXPRESA O IMPLÍCITA, INCLUYENDO PERO NO LIMITADO A GARANTÍAS DE COMERCIALIZACIÓN, IDONEIDAD PARA UN PROPÓSITO PARTICULAR E INCUMPLIMIENTO. EN NINGÚN CASO LOS AUTORES O PROPIETARIOS DE LOS DERECHOS DE AUTOR SERÁN RESPONSABLES DE NINGUNA RECLAMACIÓN, DAÑOS U OTRAS RESPONSABILIDADES, YA SEA EN UNA ACCIÓN DE CONTRATO, AGRAVIO O CUALQUIER OTRO MOTIVO, DERIVADAS DE, FUERA DE O EN CONEXIÓN CON EL SOFTWARE O SU USO U OTRO TIPO DE ACCIONES EN EL SOFTWARE. 
'''
from tkinter import *
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils import *
import re
import os

global allRow

def openFile():
    global fileName
    fileName = filedialog.askopenfilename(initialdir = ".", title = "Select file", filetypes = (("Archivos Excel","*.xlsx"), ("Cualquier archivo","*.*")))
    print(fileName)

#   Obtiene las columnas de referencias pasadas por una cadena
#   Para ser usadas con openpyxl
def getRefCols(columns_string):
    columns_string = columns_string.replace(" ", "")
    vCols = []
    if len(columns_string) > 1:
        for v in columns_string.split(','):
            vCols.append(column_index_from_string(v))
    else:
        vCols.append(column_index_from_string(columns_string))
    
    return vCols

def getIndicesAffected(str_indices):
    indicesSelected = []

    str_indices = str_indices.replace(" ", "")
    str_indices = str_indices.split(':')
    
    indicesSelected.append("".join(re.split("[^a-zA-Z]*", str_indices[0])))
    indicesSelected.append("".join(re.split("[^0-9]*", str_indices[0])))
    indicesSelected.append("".join(re.split("[^a-zA-Z]*", str_indices[1])))
    indicesSelected.append("".join(re.split("[^0-9]*", str_indices[1])))

    indicesSelected[0] = column_index_from_string(indicesSelected[0])
    indicesSelected[1] = int(indicesSelected[1])
    indicesSelected[2] = column_index_from_string(indicesSelected[2])
    indicesSelected[3] = int(indicesSelected[3])

    return indicesSelected

def removeWholeRows(row, matrixAffected):
    for cells in ws.iter_cols(
        min_col = matrixAffected[0],
        max_col = matrixAffected[2],
        min_row = row,
        max_row = row
        ):
        for cell in cells:
            cell.value = None

def getValues(refColumns, row_init, row_end):
    vRefValue = [[], []]
    for i in range(row_init, row_end + 1):
        vRefValue[1].append(i)
        str_aux = ""
        for j in refColumns:
            str_aux += str(ws[i][j - 1].value) + ", "

        vRefValue[0].append(str_aux)

    return vRefValue

def removeNextsRepeated(vRefValue, cornerMatrix, refColumns):
    currentValue = None
    i = 0
    for ref in vRefValue[0]:
        if currentValue != ref:
            currentValue = ref
            print("Nuevo valor {}".format(currentValue))
        else:
            print("A borrar, se repitió el {}".format(currentValue))
            if allRow.get():
                removeWholeRows(vRefValue[1][i], cornerMatrix)
            else:
                for col in refColumns:
                    ws[vRefValue[1][i]][col - 1].value = None
        i += 1
        
def removeUntilNextChange():
    global wb
    global ws

    wb = load_workbook(fileName)
    ws = wb[e1.get()]
    indices = getIndicesAffected(e2.get())
    refColumns = getRefCols(e3.get())
    valuesRef = getValues(refColumns, indices[1], indices[3])

    removeNextsRepeated(valuesRef, indices, refColumns)

    print("Guardando...")
    wb.save("mod.xlsx")
    print("Terminé :)")

root = Tk()
root.title("Herramientas Excel")

root.geometry("500x500")
root.resizable(0, 0)

allRow = BooleanVar()

Label(root, text = "Hoja del libro").grid(row = 0)
Label(root, text = "Seleccione la matriz\nEjemplo: D3:F5").grid(row = 1)
Label(root, text = "Columnas referentes\nSeparadas por ','").grid(row = 2)
Button(root, text = "Abrir archivo", command = openFile).grid(row = 3, column = 0, sticky = W, pady = 4)
Button(root, text = "Aceptar", command = removeUntilNextChange).grid(row = 3, column = 1, sticky = W, pady = 4)
Button(root, text = "Salir", command = root.quit).grid(row = 3, column = 2, sticky = W, pady = 4)
Checkbutton(root, text = "Aplicar eliminación a toda la fila", variable = allRow, onvalue = TRUE, offvalue = FALSE).grid(row = 1, column = 2, sticky = W, pady = 4)

e1 = Entry(root)
e2 = Entry(root)
e3 = Entry(root)

e1.grid(row = 0, column = 1)
e2.grid(row = 1, column = 1)
e3.grid(row = 2, column = 1)

mainloop()
