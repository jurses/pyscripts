#!/usr/bin/env python3
'''
Copyright 2018 Raúl Ulises Martín Hernández

Se concede permiso por la presente, libre de cargos, a cualquier persona que obtenga una copia de este software y de los archivos de documentación asociados (el "Software"), a utilizar el Software sin restricción, incluyendo sin limitación los derechos a usar, copiar, modificar, fusionar, publicar, distribuir, sublicenciar, y/o vender copias del Software, y a permitir a las personas a las que se les proporcione el Software a hacer lo mismo, sujeto a las siguientes condiciones:

El aviso de copyright anterior y este aviso de permiso se incluirán en todas las copias o partes sustanciales del Software.

EL SOFTWARE SE PROPORCIONA "COMO ESTÁ", SIN GARANTÍA DE NINGÚN TIPO, EXPRESA O IMPLÍCITA, INCLUYENDO PERO NO LIMITADO A GARANTÍAS DE COMERCIALIZACIÓN, IDONEIDAD PARA UN PROPÓSITO PARTICULAR E INCUMPLIMIENTO. EN NINGÚN CASO LOS AUTORES O PROPIETARIOS DE LOS DERECHOS DE AUTOR SERÁN RESPONSABLES DE NINGUNA RECLAMACIÓN, DAÑOS U OTRAS RESPONSABILIDADES, YA SEA EN UNA ACCIÓN DE CONTRATO, AGRAVIO O CUALQUIER OTRO MOTIVO, DERIVADAS DE, FUERA DE O EN CONEXIÓN CON EL SOFTWARE O SU USO U OTRO TIPO DE ACCIONES EN EL SOFTWARE. 
'''

from openpyxl import load_workbook
from openpyxl.utils import *
from decimal import Decimal
import re
import os
import sys

#sys.stdout.flush()

# Recibiremos:
#   Coordenadas de la matriz afectada <column><row>:<column><row>
#   Columnas de referencia
#   Eliminar todos los elementos o algunos de la fila

def getMatrixCoordinates(string_matrix_coordinate):
    indicesSelected = []

    string_matrix_coordinate = string_matrix_coordinate.replace(" ", "")
    string_matrix_coordinate = string_matrix_coordinate.split(':')
    
    indicesSelected.append("".join(re.split("[^a-zA-Z]*", string_matrix_coordinate[0])))
    indicesSelected.append("".join(re.split("[^0-9]*", string_matrix_coordinate[0])))
    indicesSelected.append("".join(re.split("[^a-zA-Z]*", string_matrix_coordinate[1])))
    indicesSelected.append("".join(re.split("[^0-9]*", string_matrix_coordinate[1])))

    indicesSelected[0] = column_index_from_string(indicesSelected[0])
    indicesSelected[1] = int(indicesSelected[1])
    indicesSelected[2] = column_index_from_string(indicesSelected[2])
    indicesSelected[3] = int(indicesSelected[3])
    
    print("Indices:", indicesSelected)

    return indicesSelected

def getIndicesReference(string_columns):
    vReferences = []

    string_columns = string_columns.replace(" ", "")
    string_columns = string_columns.split(",")

    for s in string_columns:
        vReferences.append(column_index_from_string(s))

    print("vReferencias:", vReferences)

    return vReferences

def getRowReferences(vRow):
    string_reference = ""

    for col in columns_reference:
        string_reference += str(vRow[col - matrixCoord[0]].value) + ", "

    print("Cadena de referencia de la fila: {}, {}".format(vRow[0].row, string_reference))

    return string_reference

def printPercentageCompleted(current, last):
    print("{}/{} líneas".format(current, last), flush = True)
    print("{}%".format(round(100 * current / last, 2)), flush = True)

def removeRow(row):
    for cell in row:
        cell.value = None

def removeSelective(vRow):
    for col in columns_toClear:
        vRow[col - matrixCoord[0]].value = None

def removeRepeated():
    global matrixCoord
    global reference
    global columns_reference
    global columns_toClear

    saveEachXRows = 100
    work_sheet_string = "C Tpatterns"
    columns_reference = getIndicesReference("O, T, U")
    columns_toClear = getIndicesReference("K, J, L, O, T, U")
    matrixCoord = getMatrixCoordinates("J6730:U8169")
    currentRef = ""

    allRow = True

    print("Cargando el archivo...")
    wb = load_workbook("21 VM i steps Marro1 per rols_Versión Theme con letras-REV.xlsx")
    print("Cargando la página...")
    ws = wb["C Tpatterns"]

    for i, cells in enumerate(ws.iter_rows(
        min_col = matrixCoord[0],
        max_col = matrixCoord[2],
        min_row = matrixCoord[1],
        max_row = matrixCoord[3]
        )):
        if currentRef != getRowReferences(cells):
            currentRef = getRowReferences(cells)
        else:
            if allRow:
                removeRow(cells)
            else:
                removeSelective(cells)

        printPercentageCompleted(i, matrixCoord[3] - matrixCoord[1])

    print("Guardando final...")
    wb.save("comp.21 VM i steps Marro1 per rols_Versión Theme con letras-REV.xlsx")
    print("Fin.")

removeRepeated()
