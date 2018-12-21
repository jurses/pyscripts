#!/usr/bin/env python3
'''
Copyright 2018 Raúl Ulises Martín Hernández

Se concede permiso por la presente, libre de cargos, a cualquier persona que obtenga una copia de este software y de los archivos de documentación asociados (el "Software"), a utilizar el Software sin restricción, incluyendo sin limitación los derechos a usar, copiar, modificar, fusionar, publicar, distribuir, sublicenciar, y/o vender copias del Software, y a permitir a las personas a las que se les proporcione el Software a hacer lo mismo, sujeto a las siguientes condiciones:

El aviso de copyright anterior y este aviso de permiso se incluirán en todas las copias o partes sustanciales del Software.

EL SOFTWARE SE PROPORCIONA "COMO ESTÁ", SIN GARANTÍA DE NINGÚN TIPO, EXPRESA O IMPLÍCITA, INCLUYENDO PERO NO LIMITADO A GARANTÍAS DE COMERCIALIZACIÓN, IDONEIDAD PARA UN PROPÓSITO PARTICULAR E INCUMPLIMIENTO. EN NINGÚN CASO LOS AUTORES O PROPIETARIOS DE LOS DERECHOS DE AUTOR SERÁN RESPONSABLES DE NINGUNA RECLAMACIÓN, DAÑOS U OTRAS RESPONSABILIDADES, YA SEA EN UNA ACCIÓN DE CONTRATO, AGRAVIO O CUALQUIER OTRO MOTIVO, DERIVADAS DE, FUERA DE O EN CONEXIÓN CON EL SOFTWARE O SU USO U OTRO TIPO DE ACCIONES EN EL SOFTWARE. 
'''
#from tkinter import *
#from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.utils import *
import re
import os

# filename: string, the file name.
# columns_string: string,   a string with the columns name in. Reference columns.
# vCols: vector,    filled with columns' index. Reference columns.
# str_indices: string,  range of matrix expressed this way: <col><row>:<col><row>
# indicesSelected: vector, a vector of size 4 with the previous elements. [<index row>, <index column>, <index row>, <index column>]
# vRefValue:  vector with 2 vectors in. The first vector contains the reference columns in a string from the first row to the last row 
#    of the affected matrix  and the second vector contains the rows indices.

global allRow

def selectFile():
    global fileName

    fileName = "21 VM i steps Marro1 per rols_Versión Theme con letras-REV.xlsx"
    # fileName = filedialog.askopenfilename(initialdir = ".", title = "Select file", filetypes = (("Archivos Excel","*.xlsx"), ("Cualquier archivo", "*.*")))
    print("Archivo seleccionado: {}".format(fileName))

def getRefColsCoord(columns_string):
    columns_string = columns_string.replace(" ", "")
    vCols = []
    if len(columns_string) > 1:
        for v in columns_string.split(','):
            vCols.append(column_index_from_string(v))
    else:
        vCols.append(column_index_from_string(columns_string))

    print("Las columnas seleccionadas han sido:")
    for col in vCols:
        print(get_column_letter(col))

    print("_____")
    
    return vCols

def getIndicesAffected(str_indices):
    indicesSelected = []

    str_indices = str_indices.replace(" ", "")
    str_indices = str_indices.split(':')
    
    indicesSelected.append("".join(re.split("[^a-zA-Z]*", str_indices[0])))
    indicesSelected.append("".join(re.split("[^0-9]*", str_indices[0])))
    indicesSelected.append("".join(re.split("[^a-zA-Z]*", str_indices[1])))
    indicesSelected.append("".join(re.split("[^0-9]*", str_indices[1])))

    indicesSelected[0] = column_index_from_string(indicesSelected[0])
    indicesSelected[1] = int(indicesSelected[1])
    indicesSelected[2] = column_index_from_string(indicesSelected[2])
    indicesSelected[3] = int(indicesSelected[3])

    return indicesSelected

def removeWholeRows(row, matrixAffected):
    for cells in ws.iter_cols(
        min_col = matrixAffected[0],
        max_col = matrixAffected[2],
        min_row = row,
        max_row = row
        ):
        for cell in cells:
            cell.value = None

def getValues(refColumns, row_init, row_end):
    vRefValue = [[], []]
    print("Obteniendo los elementos referentes...")
    for i in range(row_init, row_end + 1):
        vRefValue[1].append(i)
        str_aux = ""

        for j in refColumns:
            str_aux += str(ws[i][j - 1].value) + ", "

        vRefValue[0].append(str_aux)
        print("{}%".format(int((100 * (i - row_init) / (row_end - row_init)))))

    return vRefValue

def printPercentCompleted(row):
    print("{}%".format(int((100 * (row - row_init) / (row_end - row_init)))))

def removeNextsRepeated(vRefValue, cornerMatrix, refColumns):
    global row_init
    global row_end

    row_init = cornerMatrix[1]
    row_end = cornerMatrix[3]
    currentValue = None

    for i, ref in enumerate(vRefValue[0]):
        if currentValue != ref:
            currentValue = ref
            print("Fila: {}, nueva cadena: \"{}\".".format(vRefValue[1][i], currentValue))
        else:
            print("Fila: {}, se repite la cadena.".format(vRefValue[1][i]))
            if allRow:
                removeWholeRows(vRefValue[1][i], cornerMatrix)
            else:
                for col in refColumns:
                    ws[vRefValue[1][i]][col - 1].value = None

            printPercentCompleted(vRefValue[1][i])

def removeUntilNextChange():
    global wb
    global ws

    work_sheet_string = "C Tpatterns"
    #matrix_range_string = "J6730:U6771"
    matrix_range_string = "J6730:U8169"
    ref_columns_string = "O, T, U"

    #work_sheet_string = e1.get()
    #matrix_range_string = e2.get()
    #ref_columns_string = e3.get()

    selectFile()  # borrar esta llamada en GUI

    print("Abriendo el archivo...")
    wb = load_workbook(fileName)
    print("Abriendo la hoja...")
    ws = wb[work_sheet_string]
    indices = getIndicesAffected(matrix_range_string)
    refColumns = getRefColsCoord(ref_columns_string)
    valuesRef = getValues(refColumns, indices[1], indices[3])

    removeNextsRepeated(valuesRef, indices, refColumns)

    print("Guardando...")
    wb.save("mod.xlsx")
    print("Terminé :)")

#root = Tk()
#root.title("Herramientas Excel")

#root.geometry("500x500")
#root.resizable(0, 0)

allRow = True

#Label(root, text = "Hoja del libro").grid(row = 0)
#Label(root, text = "Seleccione la matriz\nEjemplo: D3:F5").grid(row = 1)
#Label(root, text = "Columnas referentes\nSeparadas por ','").grid(row = 2)
#Button(root, text = "Abrir archivo", command = selectFile).grid(row = 3, column = 0, sticky = W, pady = 4)
#Button(root, text = "Aceptar", command = removeUntilNextChange).grid(row = 3, column = 1, sticky = W, pady = 4)
#Button(root, text = "Salir", command = root.quit).grid(row = 3, column = 2, sticky = W, pady = 4)
#Checkbutton(root, text = "Aplicar eliminación a toda la fila", variable = allRow, onvalue = TRUE, offvalue = FALSE).grid(row = 1, column = 2, sticky = W, pady = 4)

#e1 = Entry(root)
#e2 = Entry(root)
#e3 = Entry(root)

#E1.grid(row = 0, column = 1)
#e2.grid(row = 1, column = 1)
#e3.grid(row = 2, column = 1)

removeUntilNextChange()
